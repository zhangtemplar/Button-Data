# coding=utf-8
__author__ = "Qiang Zhang"
__maintainer__ = "Qiang Zhang"
__email__ = "zhangtemplar@gmail.com"
"""
Add documentation of this module here.
"""

import json
import os
import re
import traceback
from copy import deepcopy
from datetime import datetime
from dateutil import parser
import tabula
from openpyxl import Workbook, load_workbook
from base.template import create_product, create_relationship, create_company, add_record
from base.util import remove_empty_string_from_array, create_logger


def extract_brand_company(text):
    elements = text.split('(')
    if len(elements) > 0:
        return '('.join(elements[0:-1]), elements[-1][:-1]
    else:
        print('Fail to extract brand name and company name from {}'.format(text))
        return text, ''


def extract_classification(text):
    for m in re.finditer(r'^([a-zA-Z\ \&\r]+[0-9\r]+)(.+)$', text):
        return m.group(1), m.group(2)
        break
    return text, ''


def replace_carriage(text):
    return text.replace('\r', ' ')


def device_page_to_json_2017(file_name, page):
    df = tabula.read_pdf(
        file_name,
        spreadsheet=True,
        output_format='json',
        pages=page)
    # note line 0 is useless
    # line 1 is the header
    result = []
    offset = 1 if len(df[0]['data']) % 2 == 1 else 2
    if len(df[0]['data']) < 2 + offset:
        return result
    header = ['Review Category', 'Approval Date', 'Approval Date in US Clinical Study Results: Japanese/Foreign',
              'Brand Name (Applicant Company)', 'New Approval /Partial Change', 'Classification Term Name',
              'Notes']
    second_header = [
        'Approval Date Detail', 'Clinical Study Results: Japanese/Foreign', 'NA', 'Classification Term Name Detail',
        'NA', 'NA', 'NA']
    for row in range(offset, len(df[0]['data']) - 1, 2):
        # remove the empty cell
        left = 0
        while left < len(df[0]['data'][row]) and len(df[0]['data'][row][left]['text']) <= 0:
            left += 1
        df[0]['data'][row] = df[0]['data'][row][left:]
        left = 0
        while left < len(df[0]['data'][row + 1]) and len(df[0]['data'][row + 1][left]['text']) <= 0:
            left += 1
        df[0]['data'][row + 1] = df[0]['data'][row + 1][left:]
        data = {k: v['text'] for k, v in zip(header, df[0]['data'][row])}
        data.update({k: v['text'] for k, v in zip(second_header, df[0]['data'][row + 1]) if k != 'NA'})
        data['Brand Name'], data['Applicant Company'] = extract_brand_company(data['Brand Name (Applicant Company)'])
        del data['Brand Name (Applicant Company)']
        result.append(data)
    return result


def device_pdf_to_json_2017(file_name, number_pages):
    result = []
    for page in range(1, number_pages + 1):
        try:
            data = device_page_to_json_2017(file_name, page)
            result.extend(data)
        except Exception as e:
            print(e)
    print('{} pages are processed'.format(page))
    return result


def device_page_to_json_pre_2016(file_name, page, header):
    df = tabula.read_pdf(
        file_name,
        spreadsheet=True,
        output_format='json',
        pages=page)
    # note line 0 is useless
    # line 1 is the header
    result = []
    offset = 1 if len(df[0]['data']) % 2 == 1 else 2
    if len(df[0]['data']) < 2 + offset:
        return result
    for row in range(offset, len(df[0]['data'])):
        # remove the empty cell
        left = 0
        while left < len(df[0]['data'][row]) and len(df[0]['data'][row][left]['text']) <= 0:
            left += 1
        right = len(df[0]['data'][row]) - 1
        while right >= 0 and len(df[0]['data'][row][right]['text']) <= 0:
            right -= 1
        df[0]['data'][row] = df[0]['data'][row][left:right + 1]
        data = {k: v['text'] for k, v in zip(header, df[0]['data'][row])}
        elements = data['Approval Date'].split('\r')
        if len(elements) > 0:
            data['Approval Date Detail'] = '\r'.join(elements[1:])
            data['Approval Date'] = elements[0]
        else:
            data['Approval Date Detail'] = ''
        if 'Approval Date in US Clinical Study Results: Japanese/Foreign' not in data:
            data['Approval Date in US Clinical Study Results: Japanese/Foreign'] = ''
        elements = data['Approval Date in US Clinical Study Results: Japanese/Foreign'].split('\r')
        if len(elements) > 0:
            data['Clinical Study Results: Japanese/Foreign'] = '\r'.join(elements[1:])
            data['Approval Date in US Clinical Study Results: Japanese/Foreign'] = elements[0]
        else:
            data['Clinical Study Results: Japanese/Foreign'] = ''
        data['Brand Name'], data['Applicant Company'] = extract_brand_company(data['Brand Name (Applicant Company)'])
        del data['Brand Name (Applicant Company)']
        data['Classification Term Name'], data['Classification Term Name Detail'] = extract_classification(
            data['Classification Generic Name'])
        del data['Classification Generic Name']
        if 'No.' in data:
            del data['No.']
        result.append(data)
    return result


def device_pdf_to_json_2008_2016(file_name, number_pages):
    result = []
    header = ['Review Category', 'Approval Date', 'Approval Date in US Clinical Study Results: Japanese/Foreign',
              'No.', 'Brand Name (Applicant Company)', 'New Approval /Partial Change', 'Classification Generic Name',
              'Notes']
    for page in range(1, number_pages + 1):
        try:
            data = device_page_to_json_pre_2016(file_name, page, header)
            result.extend(data)
        except Exception as e:
            print(e)
            break
    print('{} pages are processed'.format(page))
    return result


def device_pdf_to_json_2007(file_name, number_pages):
    result = []
    header = ['Review Category', 'Approval Date',
              'No.', 'Brand Name (Applicant Company)', 'New Approval /Partial Change', 'Classification Generic Name',
              'Notes']
    for page in range(1, number_pages + 1):
        try:
            data = device_page_to_json_pre_2016(file_name, page, header)
            result.extend(data)
        except Exception as e:
            print(e)
            break
    print('{} pages are processed'.format(page))
    return result


def device_pdf_to_json_pre_2006(file_name, number_pages):
    result = []
    header = ['Review Category', 'Approval Date', 'Brand Name (Applicant Company)', 'New Approval /Partial Change',
              'Classification Generic Name', 'Notes']
    for page in range(1, number_pages + 1):
        try:
            data = device_page_to_json_pre_2016(file_name, page, header)
            result.extend(data)
        except Exception as e:
            print(e)
            break
    print('{} pages are processed'.format(page))
    return result


def get_device(work_directory):
    device = []
    device = device_pdf_to_json_2017(
        os.path.join(work_directory, 'FY 2017 New & Improved (with Clinical Data) Medical Devices.pdf'), 16)
    for year, page in zip(range(2009, 2017), [6, 6, 12, 13, 24, 18, 20, 14]):
        print('{}/{}'.format(year, page))
        device.extend(device_pdf_to_json_2008_2016(
            os.path.join(work_directory, 'FY {} New & Improved (with Clinical Data) Medical Devices.pdf'.format(year)),
            page))
    device.extend(device_pdf_to_json_2008_2016(
        os.path.join(work_directory, 'FY 2008 New & Other (with Clinical Data) Medical Devices.pdf'), 5))
    device.extend(device_pdf_to_json_2007(
        os.path.join(work_directory, 'FY 2007 New & Other (with Clinical Data) Medical Devices.pdf'), 6))
    for year, page in zip(range(2004, 2007), [2, 1, 2]):
        print('{}/{}'.format(year, page))
        device.extend(device_pdf_to_json_pre_2006(
            os.path.join(work_directory, 'FY {} New Medical Devices.pdf'.format(year)), page))
    with open(os.path.join(work_directory, 'japan_device.json'), 'w') as fo:
        json.dump(device, fo)
    return device


def split_record(row):
    # the table may group several drugs with the same ingradient into a single cell
    # and organized as: drug1\rdrug2\r(manufacture)\rdrug1\rdrug2\(manufacture)
    status = row['New Approval /Partial Change'].split('\r')
    name = row['Brand Name (Applicant Company)'].split('\r')
    if len(status) < len(name):
        status.extend(['' for i in range(len(name) - len(status))])
    result = []
    manufacture = ''
    brand = ''
    approval = ''
    track_manufacture = False
    for left, right in zip(status, name):
        if len(left) > 1:
            # start of brand name
            if len(brand) > 0:
                result.append([brand, '', approval])
            # reset
            brand = right
            approval = left
            manufacture = ''
        else:
            if right.startswith('('):
                # insert result
                result.append([brand, '', approval])
                # start of manufacture name
                if right.endswith(')'):
                    # end of manufacture name in the same line
                    manufacture = right[1:-1]
                    # add manufacture to previous result
                    for r in result:
                        if len(r[1]) < 1:
                            r[1] = manufacture
                else:
                    # reset
                    manufacture = right[1:]
                    track_manufacture = True
                brand = ''
                approval = ''
            elif right.endswith(')'):
                # end of manufacture name
                manufacture += ' '
                manufacture += right[:-1]
                track_manufacture = False
                # add manufacture to previous result
                for r in result:
                    if len(r[1]) < 1:
                        r[1] = manufacture
            else:
                if track_manufacture:
                    # track manufacture name now
                    manufacture += ' '
                    manufacture += right
                else:
                    # continuation of brand name
                    brand += ' '
                    brand += right
    return result


def drug_page_to_json_pre_2016(file_name, page, header):
    df = tabula.read_pdf(
        file_name,
        spreadsheet=True,
        output_format='json',
        pages=page)
    # note line 0 is useless
    # line 1 is the header
    result = []
    offset = 1 if len(df[0]['data']) % 2 == 1 else 2
    if offset == 2 and page != 1:
        # header only applies to page 1
        offset = 0
    if len(df[0]['data']) < 2 + offset:
        return result
    for row in range(offset, len(df[0]['data'])):
        # determine the combined table
        number_valid_cells = 0
        for v in df[0]['data'][row]:
            if v['width'] > 0.1:
                number_valid_cells += 1
        df[0]['data'][row] = df[0]['data'][row][number_valid_cells:] + df[0]['data'][row][:number_valid_cells]
        data = {k: v['text'] for k, v in zip(header, df[0]['data'][row])}
        elements = split_record(data)
        data['Brand Name'], data['Applicant Company'] = extract_brand_company(data['Brand Name (Applicant Company)'])
        del data['Brand Name (Applicant Company)']
        if 'No.' in data:
            del data['No.']
        for e in elements:
            new_data = deepcopy(data)
            new_data['Brand Name'] = e[0]
            new_data['Applicant Company'] = e[1]
            new_data['New Approval /Partial Change'] = e[2]
            result.append(new_data)
    return result


def drug_pdf_to_json_post_2007(file_name, number_pages):
    result = []
    header = ['Review Category', 'Approval Date', 'No.', 'Brand Name (Applicant Company)',
              'New Approval /Partial Change', 'Active Ingredient (underlined: new active ingredient)',
              'Notes']
    for page in range(1, number_pages + 1):
        try:
            data = drug_page_to_json_pre_2016(file_name, page, header)
            result.extend(data)
        except Exception as e:
            traceback.print_tb(e.__traceback__)
            break
    print('{} pages are processed'.format(page))
    return result


def drug_pdf_to_json_pre_2006(file_name, number_pages):
    result = []
    header = ['Review Category', 'Approval Date', 'Brand Name (Applicant Company)',
              'New Approval /Partial Change', 'Active Ingredient (underlined: new active ingredient)',
              'Notes']
    for page in range(1, number_pages + 1):
        try:
            data = drug_page_to_json_pre_2016(file_name, page, header)
            result.extend(data)
        except Exception as e:
            traceback.print_tb(e.__traceback__)
            break
    print('{} pages are processed'.format(page))
    return result


def get_drug(work_directory):
    drug = []
    drug = drug_pdf_to_json_post_2007(os.path.join(work_directory, 'FY 2018 (April – October).pdf'), 8)
    for year, page in zip(range(2007, 2018), [7, 5, 7, 8, 15, 11, 9, 10, 10, 10, 12, 8]):
        print('{}/{}'.format(year, page))
        drug.extend(drug_pdf_to_json_post_2007(os.path.join(work_directory, 'FY {}.pdf'.format(year)), page))
    # ignore the result prior to 2006, as the data is malformed
    # for year, page in zip(range(2004, 2007), [10, 5, 7]):
    #     print('{}/{}'.format(year, page))
    #     drug.extend(drug_pdf_to_json_pre_2006('japan/FY {}.pdf'.format(year), page))
    with open(os.path.join(work_directory, 'japan_drug.json'), 'w') as fo:
        json.dump(drug, fo)
    return drug


def parse_device(cells):
    review_category = {
        '1': 'Ophthalmology and otorhinolaryngology',
        '2': 'dentistry',
        '3': 'cerebral, cardiovascular, respiratory, psychiatric, and neurological field',
        '3-1': 'Intervention devices mainly in cerebral, cardiovascular, respiratory, psychiatric, and neurological field',
        '3-2': 'Non-intervention devices mainly in cerebral, cardiovascular, respiratory, psychiatric, and neurological field',
        '4': 'cerebral, cardiovascular, respiratory, psychiatric, and neurological field',
        '5': 'gastrointestinal and urinary systems, obstetrics and gynecology',
        '6': 'orthopedic/plastic surgery and dermatology',
        '7': 'laboratory tests, in vitro diagnostics',
        '8': 'multicategory medical devices, advanced electronic medical devices, and other uncategorized medical devices',
    }
    p = create_product()
    p['name'] = cells[4]
    if isinstance(cells[3], datetime.datetime):
        p['created'] = cells[3].strftime("%a, %d %b %Y %H:%M:%S GMT")
    else:
        try:
            p['created'] = parser.parse(cells[3]).strftime("%a, %d %b %Y %H:%M:%S GMT")
        except:
            pass
    if isinstance(cells[1], datetime.datetime):
        p['updated'] = cells[1].strftime("%a, %d %b %Y %H:%M:%S GMT")
    else:
        try:
            p['updated'] = parser.parse(cells[1]).strftime("%a, %d %b %Y %H:%M:%S GMT")
        except:
            pass
    p['tag'] = remove_empty_string_from_array([cells[5], cells[6], 'Japan PMDA', 'Medical Device'])
    p['asset']['lic'] = p['tag']
    p['asset']['stat'] = 2
    p['abs'] = review_category.get(cells[10], cells[10])
    if len(p['abs']) < 1:
        p['abs'] = p['name']
    p['asset']['market'] = cells[9]
    p['addr']['country'] = 'Japan'
    p['addr']['city'] = 'Unknown'

    a = create_company()
    a['name'] = cells[0]
    a['abs'] = 'A Medical Device Company'
    a['addr'] = p['addr']
    a['tag'] = p['tag']
    return p, a


def parse_drug(cells):
    review_category = {
        '1': 'Gastrointestinal drugs, dermatologic drugs, immunosuppressive drugs, and others (not classified as other categories)',
        '2': "Cardiovascular drugs, antiparkinsonian drugs, anti-Alzheimer's drugs",
        '3-1': 'Central/peripheral nervous system drugs (excluding anesthetic drugs)',
        '3-2': 'Anesthetic drugs, sensory organ drugs (excluding drugs for inflammatory diseases), narcotics',
        '4': 'Antibacterial drugs, antiviral drugs (excluding AIDS drugs), antifungal drugs, antiprotozoal drugs, anthelmintic drugs',
        '5': 'Reproductive system drugs, drugs for urogenital system, combination drugs',
        '6-1': 'Respiratory tract drugs, anti-allergy drugs (excluding dermatologic drugs), sensory organ drugs (drugs for inflammatory diseases)',
        '6-2': 'Hormone drugs, drugs for metabolic disorders (including diabetes mellitus, osteoporosis, gout, and inborn errors of metabolism)',
    }
    p = create_product()
    p['name'] = cells[3]
    if isinstance(cells[2], datetime):
        p['created'] = cells[2]
    else:
        try:
            p['created'] = parser.parse(cells[2])
        except:
            pass
    category = review_category.get(cells[6], cells[6])
    p['tag'] = remove_empty_string_from_array([category, 'Japan PMDA', 'Drug'])
    p['asset']['lic'] = p['tag']
    p['asset']['stat'] = 2
    p['asset']['tech'] = cells[0]
    p['abs'] = cells[5]
    if len(p['abs']) < 1:
        p['abs'] = p['name']
    p['addr']['country'] = 'Japan'
    p['addr']['city'] = 'Unknown'

    a = create_company()
    a['name'] = cells[0]
    a['abs'] = 'A Drug Company'
    a['addr'] = p['addr']
    a['tag'] = p['tag']
    return p, a


def upload_to_server(work_directory):
    book = load_workbook(os.path.join(work_directory, 'japan.xlsx'))
    log = create_logger('japan-pmda')
    log.critical(datetime.datetime.now())
    for sheet_name in ('device', 'drug'):
        sheet = book.get_sheet_by_name(sheet_name)
        first_row = True
        for row in sheet.rows:
            if first_row:
                first_row = False
                continue
            cells = []
            for c in row:
                if isinstance(c.value, str):
                    cells.append(replace_carriage(c.value))
                elif c.value is None:
                    cells.append('')
                else:
                    cells.append(c.value)
            if sheet_name == 'device':
                p, a = parse_device(cells)
            else:
                p, a = parse_drug(cells)
            if len(p['name']) < 1 or len(a['name']) < 1:
                log.warning('invalid record for {}'.format(p['name']))
                continue
            response = add_record('entity', [p, a])
            if response['_status'] != 'OK':
                log.error('fail to create record for {}'.format(p['name']))
                log.error(response)
                continue
            applicant_product = create_relationship(response['_items'][1]['_id'], response['_items'][0]['_id'])
            applicant_product['type'] = 7
            applicant_product['name'] = 'Applicant'
            applicant_product['abs'] = 'Applicant'
            response = add_record('relationship', [applicant_product])
            if response['_status'] != 'OK':
                log.error('fail to create relationship for {}'.format(p['name']))
                log.error(response)
            else:
                log.debug('added {} to the system'.format(p['name']))
    log.critical(datetime.datetime.now())


def main():
    work_directory = os.path.expanduser('~/Downloads/japan')
    book = Workbook()
    device = get_device(work_directory)
    sheet = book.create_sheet(title='device')
    header = sorted(list(device[0].keys()))
    sheet.append(header)
    for r in device:
        sheet.append([r.get(k, '') for k in header])

    drug = get_drug(work_directory)
    sheet = book.create_sheet(title='drug')
    header = sorted(list(drug[0].keys()))
    sheet.append(header)
    for r in drug:
        sheet.append([r.get(k, '') for k in header])

    book.save('japan.xlsx')


if __name__ == '__main__':
    main()
